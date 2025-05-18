from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException

import time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment



def test():
    service = Service('./chromedriver')
    options = webdriver.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-setuid-sandbox")
    options.add_argument("--headless")
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(service=service, options=options)

    driver.get('https://xn--80acmlhv0b.xn--90as2d.xn--80asehdb/index.php?page=four')
    test_text = driver.find_element('xpath', '//html[1]/body[1]/div[1]/div[1]/center[1]/form[1]/div[1]')
    driver.close()
    
def test_auth(user_login, user_password):
    service = Service('./chromedriver')
    options = webdriver.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-setuid-sandbox")
    options.add_argument("--headless")
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(service=service, options=options)
    
    driver.get('http://82.116.39.87:82/overtime/logon.php')
    
    driver.find_element('xpath', '//*[@id="aut"]/input[1]').send_keys(user_login)
    driver.find_element('xpath', '//*[@id="aut"]/input[2]').send_keys(user_password)
    
    driver.find_element('xpath', '//*[@id="autoriz"]/button[1]').click()
    time.sleep(1)
    
    try:
        user_name = driver.find_element('xpath', '//html[1]/body[1]/div[1]/div[1]/table[2]/tbody[1]/tr[1]/td[1]').text
        ret = 'Вы авторизованы на сайте как ' + user_name
        driver.close()
        return ret
    except NoSuchElementException:
        driver.close()
        return 'Не удалось авторизоватся на сайте\nПопробуйта зарегистрироватся с начала либо обратитесь к администратору'
        
def get_km(user_login, user_password):
    try:
        service = Service('./chromedriver')
        options = webdriver.ChromeOptions()
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-setuid-sandbox")
        options.add_argument("--headless")
        options.add_experimental_option("detach", True)
        driver = webdriver.Chrome(service=service, options=options)
    
        driver.get('http://82.116.39.87:82/overtime/logon.php')
    
        driver.find_element('xpath', '//*[@id="aut"]/input[1]').send_keys(user_login)
        driver.find_element('xpath', '//*[@id="aut"]/input[2]').send_keys(user_password)
        driver.find_element('xpath', '//*[@id="autoriz"]/button[1]').click()
        time.sleep(0.5)
        

        driver.get('http://82.116.39.87:82/overtime/mileage.php')
        time.sleep(1)
        
        kol_iter = driver.find_elements('xpath', '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr')
        
        kol = 0
        
        for i in range(0, len(kol_iter) - 1):
            iter = i + 1
            path = '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[' + str(iter) + ']/td[6]'
            number = driver.find_element('xpath', path).text
            kol += int(number)
        
        driver.close()
        ret = 'В этом мясецу вы проехали ' + str(kol) + 'км'
        return ret
    except NoSuchElementException:
        driver.close()
        return 'Error'
        print('Error')

def get_list(user_login, user_password, chat_id):
    try:
        service = Service('./chromedriver')
        options = webdriver.ChromeOptions()
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-setuid-sandbox")
        options.add_argument("--headless")
        options.add_experimental_option("detach", True)
        driver = webdriver.Chrome(service=service, options=options)
    
        driver.get('http://82.116.39.87:82/overtime/logon.php')
    
        driver.find_element('xpath', '//*[@id="aut"]/input[1]').send_keys(user_login)
        driver.find_element('xpath', '//*[@id="aut"]/input[2]').send_keys(user_password)
        driver.find_element('xpath', '//*[@id="autoriz"]/button[1]').click()
        time.sleep(0.5)
        
        driver.get('http://82.116.39.87:82/overtime/mileage.php')
        time.sleep(1)
        
        kol_iter = driver.find_elements('xpath', '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr')
        
        kol = 2
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.column_dimensions["A"].width = 5
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        sheet.column_dimensions["B"].width = 14
        sheet["B1"].alignment = Alignment(horizontal="center", vertical="center")
        
        sheet.column_dimensions["C"].width = 12
        sheet["C1"].alignment = Alignment(horizontal="center", vertical="center")
        
        sheet.column_dimensions["D"].width = 20
        sheet["D1"].alignment = Alignment(horizontal="center", vertical="center")
        
        sheet.column_dimensions["E"].width = 100
        sheet["E1"].alignment = Alignment(horizontal="center", vertical="center")
        
        sheet.column_dimensions["F"].width = 8
        sheet["F1"].alignment = Alignment(horizontal="center", vertical="center")
        
        sheet.column_dimensions["G"].width = 13
        sheet["G1"].alignment = Alignment(horizontal="center", vertical="center")
        
        sheet.column_dimensions["H"].width = 100
        sheet["H1"].alignment = Alignment(horizontal="center", vertical="center")
        
        sheet.cell(row=1, column=1).value = "№"
        sheet.cell(row=1, column=2).value = "ID"
        sheet.cell(row=1, column=3).value = "Дата"
        sheet.cell(row=1, column=4).value = "Населённый пункт"
        sheet.cell(row=1, column=5).value = "Адрес"
        sheet.cell(row=1, column=6).value = "КМ"
        sheet.cell(row=1, column=7).value = "Куда"
        sheet.cell(row=1, column=8).value = "Комментарий"
        
        for i in range(0, len(kol_iter) - 1):
            iter = i + 1
            path_number = '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[' + str(iter) + ']/td[1]'
            number = driver.find_element('xpath', path_number).text
            sheet.cell(row=i + 2, column=1).value = number
            
            path_id = '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[' + str(iter) + ']/td[2]'
            id_ticket = driver.find_element('xpath', path_id).text
            sheet.cell(row=i + 2, column=2).value = id_ticket
            
            
            path_date = '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[' + str(iter) + ']/td[3]'
            date = driver.find_element('xpath', path_date).text
            sheet.cell(row=i + 2, column=3).value = date
            sheet.cell(row=i + 2, column=3).alignment = Alignment(horizontal="center", vertical="center")
            
            path_cyti = '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[' + str(iter) + ']/td[4]'
            cyti = driver.find_element('xpath', path_cyti).text
            sheet.cell(row=i + 2, column=4).value = cyti
            sheet.cell(row=i + 2, column=4).alignment = Alignment(horizontal="center", vertical="center")
            
            path_address = '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[' + str(iter) + ']/td[5]'
            address = driver.find_element('xpath', path_address).text
            sheet.cell(row=i + 2, column=5).value = address
            
            path_km = '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[' + str(iter) + ']/td[6]'
            km = driver.find_element('xpath', path_km).text
            sheet.cell(row=i + 2, column=6).value = int(km)
            sheet.cell(row=i + 2, column=6).alignment = Alignment(horizontal="center", vertical="center")
            
            path_kuda = '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[' + str(iter) + ']/td[7]'
            kuda = driver.find_element('xpath', path_kuda).text
            sheet.cell(row=i + 2, column=7).value = kuda
            
            path_comment = '//*[@id="user_void"]/div[1]/table[1]/tbody[2]/tr[' + str(iter) + ']/td[8]'
            comment = driver.find_element('xpath', path_comment).text
            sheet.cell(row=i + 2, column=8).value = comment
            
            kol += 1
            print(kol)
            

        driver.close()
        
        sheet.cell(row=kol, column=5).value = "Общий пробег = "
        sheet.cell(row=kol, column=5).alignment = Alignment(horizontal="right", vertical="center")
        
        iteral = '=SUM(F2:F' + str(kol -1) + ')'
        number_kol = 'F' + str(kol)
        #sheet.cell(row=kol, column=6) = iteral
        sheet[number_kol] = iteral
        sheet.cell(row=kol, column=6).alignment = Alignment(horizontal="left", vertical="center")
        
        name = str(chat_id) + '.xlsx'
        workbook.save(name)
        return name
    except NoSuchElementException:
        driver.close()
        return 'Error'
        print('Error')